from django.shortcuts import render, redirect
from .forms import CustomUserCreationForm
from django.contrib.auth import login, authenticate
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from inventory.models import Product
from django.core.exceptions import ValidationError

def home(request):
    return render(request, 'register/home.html')

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, "Inscription réussie ! Vous êtes maintenant connecté.")
            return redirect("home")
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = CustomUserCreationForm()
    return render(request, "register/register.html", {"form": form})

def custom_login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, "Connexion réussie !")
                return redirect('choice_view')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = AuthenticationForm()
    return render(request, 'register/login.html', {'form': form})

from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def add_stock_view(request):
    if request.method == 'POST':
        print("Requête POST reçue")  # Vérifie que la requête POST est reçue

        # Récupérer les données du formulaire
        codes = request.POST.getlist('code[]')
        articles = request.POST.getlist('article[]')
        dates = request.POST.getlist('date[]')
        quantities = request.POST.getlist('quantity[]')
        purchase_prices = request.POST.getlist('purchase_price[]')
        selling_prices = request.POST.getlist('selling_price[]')
        supplier_bl_numbers = request.POST.getlist('supplier_bl_number[]')
        descriptions = request.POST.getlist('description[]')

        # Loguer les données envoyées pour vérifier
        print("Données reçues : ")
        for field, values in zip(['Code', 'Article', 'Date', 'Quantité', 'Prix d\'Achat', 'Prix de Vente', 'N° BL Fournisseur', 'Description'], [codes, articles, dates, quantities, purchase_prices, selling_prices, supplier_bl_numbers, descriptions]):
            print(f"{field}: {values}")

        # Regrouper les données
        products_data = zip(codes, articles, dates, quantities, purchase_prices, 
                            selling_prices, supplier_bl_numbers, descriptions)

        # Valider chaque produit et essayer de l'ajouter à la base de données
        for product in products_data:
            code, article, date, quantity, purchase_price, selling_price, supplier_bl_number, description = product
            
            # Assurez-vous que quantity, purchase_price, selling_price sont bien des nombres
            try:
                new_product = Product(
                    code=code,
                    article=article,
                    date=date,
                    quantity=int(quantity) if quantity else 0,  # Valeur par défaut
                    purchase_price=float(purchase_price) if purchase_price else 0.0,  # Valeur par défaut
                    selling_price=float(selling_price) if selling_price else 0.0,  # Valeur par défaut
                    supplier_bl_number=supplier_bl_number,
                    description=description
                )
                
                new_product.full_clean()  # Valide l'objet
                new_product.save()  # Sauvegarde dans la base de données si valide
                print(f"Produit {code} sauvegardé")  # Confirmation de la sauvegarde
                messages.success(request, f"Produit {code} ajouté avec succès!")
            except ValidationError as e:
                print(f"Erreur lors de la sauvegarde du produit {code}: {e}")
                messages.error(request, f"Erreur lors de l'ajout du produit {code}: {e}")

        return render(request, 'register/add_stock.html')  # Restez sur la même page

    return render(request, 'register/add_stock.html')



def choice_view(request):
    return render(request, 'register/choice.html')

def stock_view(request):
    return render(request, 'register/stock.html')

@csrf_exempt
def client_view(request):
    # Render the form
    if request.method == 'GET':
        return render(request, 'register/client.html')

    # Process data and generate Excel file
    elif request.method == 'POST':
        # Parse form data from request
        table_data = {
            'Date de vente': request.POST.getlist('date_vente[]'),
            'num BL': request.POST.getlist('num_bl[]'),
            'Valeur BL': request.POST.getlist('val_bl[]'),
            'Nom client': request.POST.getlist('nom_client[]'),
            'Article': request.POST.getlist('article[]'),
            'Quantite': request.POST.getlist('quantite[]'),
            'Date de paiement': request.POST.getlist('date_paiement[]'),
            'Mode de paiement': request.POST.getlist('mode_paiement[]'),
            'Total a payer': request.POST.getlist('total_a_payer[]'),
            'Reste a payer': request.POST.getlist('reste_a_payer[]'),
            'Avoir sur le bl': request.POST.getlist('avoir_bl[]'),
            'rq1': request.POST.getlist('rq1[]'),
            'rq2': request.POST.getlist('rq2[]')
        }

        # Create DataFrame from the table data
        df = pd.DataFrame(table_data)

        # Generate Excel file in memory
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='VenteData')

        # Rewind buffer and prepare the response
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="vente_data.xlsx"'
        return response

def view_stock_view(request):
    query = request.GET.get('q', '')
    if query:
        products = Product.objects.filter(article__icontains=query)  # Recherche insensible à la casse
    else:
        products = Product.objects.all()  # Récupérer tous les produits si aucune recherche

    return render(request, 'register/view_stock.html', {'products': products, 'query': query})

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse
from io import BytesIO
from inventory.models import Product  # Assurez-vous d'importer votre modèle Product

def export_stock_to_excel(request):
    # Créer un nouveau classeur
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Stock'

    # Définir les en-têtes
    headers = ['Code', 'Article', 'Date', 'Quantité', 'Prix d\'Achat', 'Prix de Vente', 'N° BL Fournisseur', 'Description']
    worksheet.append(headers)

    # Appliquer des styles aux en-têtes
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Couleur jaune
    header_font = Font(bold=True)
    
    for cell in worksheet[1]:  # 1 correspond à la première ligne (les en-têtes)
        cell.fill = header_fill
        cell.font = header_font

    # Ajouter les données des produits
    products = Product.objects.all().values_list('code', 'article', 'date', 'quantity', 'purchase_price', 'selling_price', 'supplier_bl_number', 'description')
    
    # Définir le style des bordures
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Alternance des couleurs pour les lignes
    for row_idx, product in enumerate(products, start=2):  # Commencer à la ligne 2
        worksheet.append(product)
        
        # Appliquer les bordures
        for cell in worksheet[row_idx]:
            cell.border = thin_border
            
        # Appliquer une couleur de fond alternée
        if row_idx % 2 == 0:  # Ligne paire
            fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # Vert clair
        else:  # Ligne impaire
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanc
        
        for cell in worksheet[row_idx]:
            cell.fill = fill

    # Ajuster la largeur des colonnes
    column_widths = [15, 25, 15, 10, 15, 15, 20, 30]  # Largeurs souhaitées pour chaque colonne
    for i, width in enumerate(column_widths):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

    # Formater la colonne de date
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=3):  # Colonne de date
        for cell in row:
            if isinstance(cell.value, str):  # Vérifier si la valeur est une chaîne
                cell.number_format = 'YYYY-MM-DD'  # Format de date

    # Utiliser BytesIO pour créer une réponse HTTP avec le contenu du classeur
    output = BytesIO()
    workbook.save(output)
    output.seek(0)  # Rewind the buffer

    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock.xlsx'
    
    return response

def export_to_excel(request):
    if request.method == 'POST':
        # Create a new workbook and active sheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Tableau de Vente'

        # Define table headers
        headers = [
            'Date de vente', 'num BL', 'Valeur BL', 'Nom client', 'Article',
            'Quantite', 'Date de paiement', 'Mode de paiement', 'Total a payer',
            'Reste a payer', 'Avoir sur le bl', 'rq1', 'rq2'
        ]
        worksheet.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Collect data from form inputs
        rows = zip(
            request.POST.getlist('date_vente[]'), request.POST.getlist('num_bl[]'),
            request.POST.getlist('val_bl[]'), request.POST.getlist('nom_client[]'),
            request.POST.getlist('article[]'), request.POST.getlist('quantite[]'),
            request.POST.getlist('date_paiement[]'), request.POST.getlist('mode_paiement[]'),
            request.POST.getlist('total_a_payer[]'), request.POST.getlist('reste_a_payer[]'),
            request.POST.getlist('avoir_bl[]'), request.POST.getlist('rq1[]'), request.POST.getlist('rq2[]')
        )

        # Define border style for cells
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Append rows and apply alternating row colors
        for row_idx, row_data in enumerate(rows, start=2):
            worksheet.append(row_data)
            for cell in worksheet[row_idx]:
                cell.border = thin_border
                cell.fill = PatternFill(start_color="D9EAD3" if row_idx % 2 == 0 else "FFFFFF", fill_type="solid")  # Alternating row colors

        # Set column widths
        column_widths = [15, 10, 10, 20, 15, 10, 15, 15, 15, 15, 20, 10, 10]
        for i, width in enumerate(column_widths):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = width

        # Save workbook to in-memory file and send it as response
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=vente_data.xlsx'
        return response
    return HttpResponse("Invalid request method", status=405)